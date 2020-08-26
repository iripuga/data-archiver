#!/usr/bin/python
'''
@author: Irinej Papuga

Program za beleženje metapodatkov udeležencev raziskave.
'''
# define measurements directory and database
path = "/Users/iripuga/Documents/0.UNI/2.Magisterij/4.magistrska/data-archiver/"   ### CHANGE PATH
fileop = '/OsebniPodatki.xlsx'
filekt = '/KodirnaTabela.xlsx'
OsebniPodatki = path + fileop
KodirnaTabela = path + filekt
direkt = 'makeit' # only phrase 'makeit' creates a new directory

# importing libraries 
import os
from random import randint
import sys
from PyQt5.QtWidgets import * 
import csv
import openpyxl as pyxl # knjižnica za upravljanje s .xlsx datotekami
import json

# Do I want to generate header in .csv file, and a new directory structure
head = 'n' # input('Generate header in .csv (y/n)? ')
if head == 'y':
    WriteHeader = 1
else:
    WriteHeader = 0
  
# Check with user before starting registration
run = input("Is file 'KodirnaTabela.xlsx' UNLOCKED (y/n)? ") # če je zaklenjena ne morem pravilno not pisat
os.chdir(path) # to be sure I am on the right path :P
participant = {}

def genID():
    #check path for already existing IDs to avoid duplicates
    datadir = os.listdir()
    forbiddenID = [] # these already exist
    for el in datadir:
        if el[0] == '#':
            num = int(el[1::])
            forbiddenID.append(num)
        
    while True:
        # generating 3-digit person ID
        uniqueID = randint(100, 500) # predvidevam, da ne bo več kot 400 udeležencev v raziskavi
        if uniqueID not in forbiddenID:
            break

    return str(uniqueID)

def bmi(weight, height):
    # Izračun BMI
    h = float(height) / 100 # enota: m
    w = float(weight)

    return str(round(w / h**2, 2))

def dict2list(d):
    # dict data into list, representing csv row
    l = [d['name'], d['ID'], d['age'], d['sex'], d['height'], d['weight'], d['bmi'], d['sport'], d['chronic_medical_condition'], d['medication'], d['comments']] # ne da se iterirat v for zanki :(

    return l

# functioins that write into .xlsx
def write2xlsx(workbook, content, filename):
    if WriteHeader == 1: raise NotImplementedError('Header writting not yet available for .xlsx')

    # write list of data into xlsx row
    sheet = workbook['Sheet1']
    row = sheet.max_row + 1 # pišem v novo vrstico
    col = 1
    for value in content:
        sheet.cell(row=row, column=col).value = value
        col = col + 1

    return f'Data written in {filename}, row {str(row)}'

def anonymous(workbook, content):
    # write list of data into xlsx row
    sheet = workbook['Sheet1']

    return None

# creating a class 
# that inherits the QDialog class 
class Window(QDialog): 
  
    # constructor 
    def __init__(self): 
        super(Window, self).__init__() 
  
        # setting window title 
        self.setWindowTitle("Registracija") 
  
        # setting geometry to the window 
        self.setGeometry(100, 100, 500, 400) 
  
        # creating a group box 
        self.formGroupBox = QGroupBox("Udeleženec") 
  
        # Create boxes
        # creating spin box to select age, height, weight
        self.ageSpinBar = QSpinBox() 

        # creating combo box to select sex, medical_condition, medication
        # vprašalnik za udeležence - vprašanja 5 do 12 so razporejena po smiselnosti na registracijski in merilni obrazec
        self.sexComboBox = QComboBox() 
        self.conditionComboBox = QComboBox() 
        self.sportComboBox = QComboBox() 
        self.medsComboBox = QComboBox()
        
        # Dodam možne odgovore
        self.sexComboBox.addItems(["Moski", "Zenski", "Ne-binarni", "Ostalo"]) 
        self.conditionComboBox.addItems(["Ne", "Da"]) # Glej prilogo 6 za etično, 5. vprašanje
        self.sportComboBox.addItems(["Ne", "Da"]) # 3x/week for 1h
        self.medsComboBox.addItems(["Ne", "Da"])
  
        # creating a line edit
        self.fullnameLineEdit = QLineEdit() 
        self.heightLineEdit = QLineEdit() 
        self.weightLineEdit = QLineEdit() 
        self.commentsLineEdit = QLineEdit()

        # set buttons and form
        self.createForm() # calling the method that create the form 
        self.buttonBox = QDialogButtonBox(QDialogButtonBox.Ok | QDialogButtonBox.Cancel) # creating a dialog button for ok and cancel 
        self.buttonBox.accepted.connect(self.getInfo)  # adding action when form is accepted 
        self.buttonBox.rejected.connect(self.reject) # addding action when form is rejected 
  
        # setting lay out 
        mainLayout = QVBoxLayout() # creating a vertical layout 
        mainLayout.addWidget(self.formGroupBox) # adding form group box to the layout 
        mainLayout.addWidget(self.buttonBox) # adding button box to the layout 
        self.setLayout(mainLayout) 
  
    # get info method called when form is accepted 
    def getInfo(self): 
        # generating 3-digit person ID
        personID = genID() # predvidevam, da ne bo več kot 400 udeležencev v raziskavi

        # write data into dict format
        participant['name'] = self.fullnameLineEdit.text()
        participant['ID'] = personID
        participant['height'] = self.heightLineEdit.text()
        participant['weight'] = self.weightLineEdit.text()
        participant['bmi'] = bmi(self.weightLineEdit.text(), self.heightLineEdit.text())
        participant['age'] = self.ageSpinBar.text()
        participant['sex'] = self.sexComboBox.currentText()
        participant['sport'] = self.sportComboBox.currentText()
        participant['chronic_medical_condition'] = self.conditionComboBox.currentText()
        participant['medication'] = self.medsComboBox.currentText()
        participant['comments'] = self.commentsLineEdit.text()

        ##### write information to .xlsx table and use personID alias table for anonymization
        wbOP = pyxl.load_workbook(OsebniPodatki)
        wbKT = pyxl.load_workbook(KodirnaTabela)

        row = dict2list(participant)
        print(write2xlsx(wbOP, row[1::], fileop)) # zapišem vse razen Imena
        print(write2xlsx(wbKT, row[0:2], filekt)) # zapišem ID in Ime

        #### save with same filenames
        wbOP.save(OsebniPodatki)
        wbKT.save(KodirnaTabela)

        # generate anonymous directory structure based on personID
        directory = path + '/#' + str(personID)
        if direkt == 'makeit':
            os.mkdir(directory)  
            print(f'>> Created new directory {directory}') 
        else:
            print(f'>> FAKED creating new directory {directory}') 
        

        # printing the form information 
        dots = '' # Kako anonimno :P
        for letter in self.fullnameLineEdit.text():
            if letter not in [' ']:
                dots = dots + '*'
            else:
                dots = dots + ' '
        
        print("\n>> Data preview:")
        print("Name : {0}".format(dots))
        print("ID : {0}".format(personID))
        print("Sex : {0}".format(self.sexComboBox.currentText())) 
        print("Age : {0}".format(self.ageSpinBar.text())) 
        print("Height : {0}".format(self.heightLineEdit.text()))
        print("Weight : {0}".format(self.weightLineEdit.text()))
        print("Condition : {0}".format(self.conditionComboBox.currentText()))
        print("Medication : {0}".format(self.medsComboBox.currentText()))
        print("Sport : {0}".format(self.sportComboBox.currentText()))
        print("Comments : {0}".format(self.commentsLineEdit.text()))

        # closing the window 
        self.close() 
  
    # create form method 
    def createForm(self): 
  
        # creating a form layout 
        layout = QFormLayout() 
  
        # adding rows to fomr layout
        layout.addRow(QLabel("Ime in Priimek"), self.fullnameLineEdit) 
        layout.addRow(QLabel("Spol"), self.sexComboBox)  
        layout.addRow(QLabel("Starost"), self.ageSpinBar) 
        layout.addRow(QLabel("Višina / cm"), self.heightLineEdit)
        layout.addRow(QLabel("Teža / kg"), self.weightLineEdit)
        layout.addRow(QLabel("Šport 3x na teden vsaj 1h?"), self.sportComboBox)
        layout.addRow(QLabel("Ali imate katero od navedenih zdravstvenih težav? \n- bolezni srca in ožilja, \n- visok krvni pritisk, \n- motnje srčnega ritma, \n- preobčutljivost za mraz, \n- motnje krvnega obtoka, \n- Raynaudova bolezen"), self.conditionComboBox)
        layout.addRow(QLabel("Ali redno jemljete kakšna zdravila? \n(Napišite v opombe katera)"), self.medsComboBox)
        layout.addRow(QLabel("Opombe"), self.commentsLineEdit)
  
        # setting layout 
        self.formGroupBox.setLayout(layout) 


if run == 'y':
    # main method 
    if __name__ == '__main__': 
        app = QApplication(sys.argv) # create pyqt5 app 
        window = Window() # create the instance of our Window 
        window.show()  # showing the window 
        sys.exit(app.exec()) # start the app 
else:
    raise ValueError('Decrypt file KodirnaTabela.csv before continuing!!!')

