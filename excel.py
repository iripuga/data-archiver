'''
vir: https://stackoverflow.com/questions/38022247/python-simulating-csv-dictreader-with-openpyxl#38028439
'''
# knjižnica za branje podatkov iz excela

__all__ = ['DictReader']

from openpyxl import load_workbook
from openpyxl.cell import Cell

Cell.__init__.__defaults__ = (None, None, '', None)   # Change the default value for the Cell from None to `` the same way as in csv.DictReader


class DictReader(object):
    def __init__(self, f, sheet_index,
                 fieldnames=None, restkey=None, restval=None):
        self._fieldnames = fieldnames   # list of keys for the dict
        self.restkey  = restkey         # key to catch long rows
        self.restval  = restval         # default value for short rows
        self.reader   = load_workbook(f, data_only=True).worksheets[sheet_index].iter_rows(values_only=True)
        self.line_num = 0

    def __iter__(self):
        return self

    @property
    def fieldnames(self):
        if self._fieldnames is None:
            try:
                self._fieldnames = next(self.reader)
                self.line_num += 1
            except StopIteration:
                pass

        return self._fieldnames

    @fieldnames.setter
    def fieldnames(self, value):
        self._fieldnames = value

    def __next__(self):
        if self.line_num == 0:
            # Used only for its side effect.
            self.fieldnames

        row = next(self.reader)
        self.line_num += 1

        # unlike the basic reader, we prefer not to return blanks,
        # because we will typically wind up with a dict full of None
        # values
        while row == ():
            row = next(self.reader)

        d = dict(zip(self.fieldnames, row))
        lf = len(self.fieldnames)
        lr = len(row)

        if lf < lr:
            d[self.restkey] = row[lf:]
        elif lf > lr:
            for key in self.fieldnames[lr:]:
                d[key] = self.restval

        return d